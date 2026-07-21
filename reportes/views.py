"""
Módulo de Vistas para la aplicación de Reportes.
Contiene la lógica de negocio y procesamiento de peticiones HTTP/HTTPS,
incluyendo la gestión de reportes, autenticación de usuarios, perfiles,
generación de PDFs y notificaciones en el sistema.
"""

# Importaciones de Django para redirecciones, renderizado y consulta de base de datos
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.conf import settings
from django.db.models import Q
from django.core.mail import send_mail
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie

# Librerías estándar de Python
import os
import random

# Librerías de ReportLab para la generación y maquetado de reportes en PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

# Formularios y modelos locales de la aplicación
from .forms import ReporteForm
from .models import (
    Categoria,
    PerfilUsuario,
    Reporte,
    Evidencia,
    Notificacion,
    HistorialReporte,
    Resena,
)


def crear_catalogos_iniciales():
    """
    Crea las categorías por defecto en el sistema ('Basura', 'Baches en Vía Pública',
    'Fallos de electricidad', 'Fugas de agua', 'Alumbrado público', 'Semáforo dañado')
    si no existen todavía en la base de datos.
    """
    categorias_defecto = [
        ('Basura', 'BS'),
        ('Baches en Vía Pública', 'BV'),
        ('Fallos de electricidad', 'FE'),
        ('Fugas de agua', 'FA'),
        ('Alumbrado público', 'AP'),
        ('Semáforo dañado', 'SD'),
        ('Otro', 'OT'),
    ]

    for nombre, codigo in categorias_defecto:
        cat, created = Categoria.objects.get_or_create(
            nombre=nombre,
            defaults={'codigo': codigo}
        )
        if cat.codigo != codigo:
            cat.codigo = codigo
            cat.save()


def obtener_rol(usuario):
    """
    Retorna el rol correspondiente del usuario.
    Si es un superusuario de Django, retorna 'administrador'.
    Si el usuario no cuenta con un PerfilUsuario creado, lo crea automáticamente
    con el rol por defecto 'ciudadano'.
    """
    if usuario.is_superuser:
        return 'administrador'

    try:
        return usuario.perfilusuario.rol
    except PerfilUsuario.DoesNotExist:
        PerfilUsuario.objects.create(usuario=usuario, rol='ciudadano')
        return 'ciudadano'


def redireccion_por_rol(usuario):
    """
    Determina a qué panel/dashboard redirigir al usuario
    tras autenticarse de acuerdo a su nivel de privilegios (rol).
    """
    rol = obtener_rol(usuario)

    if rol in ['administrador', 'moderador', 'supervisor']:
        return redirect('panel_administrador')

    return redirect('panel_ciudadano')


def rol_requerido(*roles_permitidos):
    """
    Decorador personalizado para proteger vistas basándose en roles específicos.
    Permite el paso libre si el usuario es superusuario de Django o si
    su rol está dentro del listado de 'roles_permitidos'.
    """
    def decorador(vista):
        @login_required
        def wrapper(request, *args, **kwargs):
            rol = obtener_rol(request.user)

            if request.user.is_superuser or rol in roles_permitidos:
                return vista(request, *args, **kwargs)

            messages.error(request, 'No tienes permiso para acceder a esta sección.')
            return redirect('panel_ciudadano')

        return wrapper
    return decorador


def crear_notificacion(usuario, reporte, titulo, mensaje):
    Notificacion.objects.create(
        usuario=usuario,
        reporte=reporte,
        titulo=titulo,
        mensaje=mensaje
    )


def registrar_historial(reporte, usuario, accion, descripcion=''):
    HistorialReporte.objects.create(
        reporte=reporte,
        usuario=usuario,
        accion=accion,
        descripcion=descripcion
    )


@ensure_csrf_cookie
def login_usuario(request):
    crear_catalogos_iniciales()

    if request.user.is_authenticated:
        return redireccion_por_rol(request.user)

    mensaje = ''

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        usuario = authenticate(request, username=username, password=password)

        if usuario is not None:
            login(request, usuario)
            return redireccion_por_rol(usuario)

        # Si authenticate devuelve None, puede ser por credenciales incorrectas
        # o porque el usuario está inactivo (cuenta suspendida/bloqueada).
        try:
            usuario_db = User.objects.get(username=username)
            if not usuario_db.is_active and usuario_db.check_password(password):
                mensaje = 'Tu cuenta ha sido suspendida.'
            else:
                mensaje = 'Usuario o contraseña incorrectos'
        except User.DoesNotExist:
            mensaje = 'Usuario o contraseña incorrectos'

    return render(request, 'reportes/login.html', {'mensaje': mensaje})


@ensure_csrf_cookie
def registro_usuario(request):
    """
    Vista de Registro de Usuarios.
    Procesa el formulario de registro, valida que los campos obligatorios
    estén presentes, crea el usuario desactivado (is_active=False), genera un
    código de verificación de 6 dígitos, lo almacena en la sesión del usuario
    y envía un correo con dicho código.
    """
    crear_catalogos_iniciales()

    # Si el usuario ya está autenticado, se le redirige a su panel
    if request.user.is_authenticated:
        return redirect('panel_ciudadano')

    mensaje = ''

    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        # Validaciones de campos obligatorios y formato
        if not username or not first_name or not last_name or not email or not password1 or not password2:
            mensaje = 'Todos los campos marcados con asterisco (*) son obligatorios.'

        elif password1 != password2:
            mensaje = 'Las contraseñas no coinciden.'

        elif len(password1) < 8:
            mensaje = 'La contraseña debe tener al menos 8 caracteres.'

        elif not password1.isalnum():
            mensaje = 'La contraseña solo debe contener caracteres alfanuméricos (letras y números sin caracteres especiales).'

        elif not (any(c.isalpha() for c in password1) and any(c.isdigit() for c in password1)):
            mensaje = 'La contraseña debe contener al menos una letra y un número.'

        elif telefono and not telefono.isdigit():
            mensaje = 'El número de teléfono solo debe contener números.'

        elif User.objects.filter(username=username).exists():
            mensaje = 'Ese nombre de usuario ya existe.'

        elif User.objects.filter(email=email).exists():
            mensaje = 'Ese correo ya está registrado.'

        else:
            # Crear el usuario en estado inactivo (is_active=False)
            usuario = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name
            )
            usuario.is_active = False
            usuario.save()

            # Crear o actualizar el perfil de usuario asociado
            perfil, _ = PerfilUsuario.objects.get_or_create(usuario=usuario)
            perfil.rol = 'ciudadano'
            perfil.telefono = telefono
            perfil.save()

            # Generar código aleatorio de 6 dígitos
            codigo = str(random.randint(100000, 999999))

            # Almacenar en la sesión del usuario de forma segura
            request.session['codigo_verificacion'] = codigo
            request.session['usuario_verificacion_id'] = usuario.id

            # Cuerpo del correo con el código de verificación
            asunto = 'Código de Verificación - Registro de Usuario'
            mensaje_correo = f"""¡Hola, {first_name}!

Gracias por registrarte en nuestro Sistema de Reportes.

Para poder activar tu cuenta y acceder al sistema, por favor introduce el siguiente código de verificación de 6 dígitos en la pantalla de verificación:

👉 {codigo} 👈

Si tienes problemas, por favor vuelve a registrarte o solicita soporte al administrador.

Atentamente,
El equipo del Sistema de Reportes."""

            # Envío de correo electrónico usando SMTP configurado de Django
            try:
                send_mail(
                    asunto,
                    mensaje_correo,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False
                )
                messages.success(request, 'Se ha enviado un código de verificación a tu correo.')
            except Exception as e:
                # Fallback en caso de que falle SMTP para que no rompa el flujo
                messages.warning(request, 'No se pudo enviar el correo de verificación. Por favor contacta al soporte.')

            # Redirigir a la pantalla de verificación de correo
            return redirect('verificar_correo')

    return render(request, 'reportes/registro.html', {'mensaje': mensaje})


@ensure_csrf_cookie
def verificar_correo(request):
    """
    Vista de Verificación de Correo.
    Compara el código introducido por el usuario con el código almacenado en la sesión.
    Si coincide, activa la cuenta (is_active=True), realiza el inicio de sesión automático
    e ingresa la notificación del nuevo ciudadano para administradores y moderadores.
    """
    # Si el usuario ya está autenticado y activo, lo redirigimos a su panel correspondiente
    if request.user.is_authenticated and request.user.is_active:
        return redireccion_por_rol(request.user)

    # Recuperar datos de verificación de la sesión
    usuario_id = request.session.get('usuario_verificacion_id')
    codigo_sesion = request.session.get('codigo_verificacion')

    # Si no hay sesión de verificación activa, lo redirigimos al registro
    if not usuario_id or not codigo_sesion:
        messages.error(request, 'Sesión de verificación expirada. Por favor regístrate nuevamente.')
        return redirect('registro')

    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == 'POST':
        codigo_ingresado = request.POST.get('codigo', '').strip()

        if not codigo_ingresado:
            messages.error(request, 'El código de verificación no puede estar vacío.')
        elif codigo_ingresado == codigo_sesion:
            # Activar el usuario en la base de datos
            usuario.is_active = True
            usuario.save()

            # Notificar a administradores y moderadores sobre el registro del nuevo usuario verificado
            admins_y_mods = User.objects.filter(
                Q(is_superuser=True) | Q(perfilusuario__rol__in=['administrador', 'moderador'])
            ).distinct()

            for admin_mod in admins_y_mods:
                crear_notificacion(
                    usuario=admin_mod,
                    reporte=None,
                    titulo='Nuevo usuario verificado',
                    mensaje=f'El usuario {usuario.username} ({usuario.first_name} {usuario.last_name}) ha verificado su correo y su cuenta está activa.'
                )

            # Inicio de sesión automático
            login(request, usuario)

            # Limpiar variables de sesión de verificación
            request.session.pop('codigo_verificacion', None)
            request.session.pop('usuario_verificacion_id', None)

            messages.success(request, '¡Felicidades! Tu cuenta ha sido verificada y activada con éxito.')
            return redireccion_por_rol(usuario)
        else:
            messages.error(request, 'El código ingresado es incorrecto. Inténtalo de nuevo.')

    return render(request, 'reportes/verificar_correo.html', {'email': usuario.email})


def verificar_usuario(request):
    """
    Vista que verifica si un nombre de usuario ya existe en la base de datos.
    Retorna un JSON con la clave 'existe'.
    """
    from django.http import JsonResponse
    username = request.GET.get('username', '').strip()
    existe = False
    if username:
        existe = User.objects.filter(username=username).exists()
    return JsonResponse({'existe': existe})


def politica_privacidad(request):
    return render(request, 'reportes/politica_privacidad.html')


def cerrar_sesion(request):
    logout(request)
    return redirect('inicio')


@login_required
def dashboard(request):
    return redireccion_por_rol(request.user)


@login_required
def panel_ciudadano(request):
    reportes_base = Reporte.objects.filter(
        usuario=request.user
    ).order_by('-fecha_reporte')

    estado = request.GET.get('estado')
    buscar = request.GET.get('buscar')

    reportes = reportes_base

    if estado and estado != 'Todos':
        reportes = reportes.filter(estado=estado)

    if buscar:
        reportes = reportes.filter(
            Q(folio__icontains=buscar) |
            Q(titulo__icontains=buscar) |
            Q(descripcion__icontains=buscar) |
            Q(categoria__nombre__icontains=buscar) |
            Q(colonia__icontains=buscar)
        )

    puntos_mapa = []

    for reporte in reportes_base.exclude(latitud__isnull=True).exclude(longitud__isnull=True):
        puntos_mapa.append({
            'id': reporte.id,
            'folio': reporte.folio or 'Sin folio',
            'titulo': reporte.titulo or 'Reporte ciudadano',
            'categoria': reporte.categoria.nombre,
            'codigo_categoria': reporte.categoria.codigo,
            'estado': reporte.estado,
            'prioridad': reporte.prioridad,
            'colonia': reporte.colonia or 'Sin colonia',
            'latitud': float(reporte.latitud),
            'longitud': float(reporte.longitud),
            'url': f'/reporte/{reporte.id}/',
        })

    context = {
        'reportes': reportes,
        'total_reportes': reportes_base.count(),
        'pendientes': reportes_base.filter(estado='Pendiente').count(),
        'en_proceso': reportes_base.filter(estado='En proceso').count(),
        'resueltos': reportes_base.filter(estado='Resuelto').count(),
        'cancelados': reportes_base.filter(estado='Cancelado').count(),
        'estado_actual': estado or 'Todos',
        'buscar': buscar or '',
        'puntos_mapa': puntos_mapa,
    }

    return render(request, 'reportes/panel_ciudadano.html', context)



@login_required
def perfil(request):
    reportes_count = Reporte.objects.filter(usuario=request.user).count()

    return render(request, 'reportes/perfil.html', {
        'reportes_count': reportes_count
    })


@login_required
def editar_perfil(request):
    perfil, _ = PerfilUsuario.objects.get_or_create(usuario=request.user)

    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        password = request.POST.get('password')

        if telefono and not telefono.isdigit():
            messages.error(request, 'El número de teléfono solo debe contener números.')
            return render(request, 'reportes/editar_perfil.html', {
                'perfil': perfil
            })

        request.user.username = username
        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.email = email

        if password:
            request.user.set_password(password)

        request.user.save()

        perfil.telefono = telefono
        perfil.save()

        messages.success(request, 'Perfil actualizado correctamente.')

        if password:
            return redirect('login')

        return redirect('perfil')

    return render(request, 'reportes/editar_perfil.html', {
        'perfil': perfil
    })


@login_required
def crear_reporte(request):
    crear_catalogos_iniciales()

    if request.method == 'POST':
        form = ReporteForm(request.POST, request.FILES)

        # Capturar lista de fotografías subidas
        fotos = [f for f in request.FILES.getlist('foto') if f.name]
        if not fotos and 'foto' in request.FILES and request.FILES['foto'].name:
            fotos = [request.FILES['foto']]

        # Validaciones de cantidad y formato JPG para evidencias
        errores_foto = []
        if len(fotos) < 1:
            errores_foto.append('Es obligatorio adjuntar al menos 1 fotografía JPG de evidencia.')
        elif len(fotos) > 3:
            errores_foto.append('Únicamente se permite adjuntar un máximo de 3 fotografías.')

        for f in fotos:
            ext = f.name.split('.')[-1].lower() if '.' in f.name else ''
            if ext not in ['jpg', 'jpeg']:
                errores_foto.append(f'El archivo "{f.name}" no está en formato JPG (.jpg / .jpeg).')

        nombre_incidencia_otro = request.POST.get('nombre_incidencia_otro', '').strip()

        # Si Django generó un error genérico en 'foto' (ej. "No se envió un archivo..."), lo reemplazamos por el nuestro
        if 'foto' in form.errors:
            del form.errors['foto']

        if errores_foto:
            for err in errores_foto:
                form.add_error('foto', err)

        # Validación del campo de texto cuando se selecciona 'Otro'
        if form.is_valid():
            cat_seleccionada = form.cleaned_data.get('categoria')
            if cat_seleccionada and 'otro' in cat_seleccionada.nombre.lower() and not nombre_incidencia_otro:
                form.add_error('categoria', 'Por favor especifica el nombre de tu incidencia en el campo de texto.')

        if form.is_valid() and not errores_foto:
            try:
                reporte = form.save(commit=False)
                reporte.usuario = request.user
                reporte.estado = 'Pendiente'
                reporte.prioridad = 'Media'
                reporte.municipio = 'Acapulco de Juárez'
                reporte.foto = fotos[0]  # Primera fotografía como foto principal

                # Si seleccionó 'Otro' y especificó un nombre personalizado
                if reporte.categoria and 'otro' in reporte.categoria.nombre.lower() and nombre_incidencia_otro:
                    nombre_custom = nombre_incidencia_otro.strip().capitalize()
                    cat_custom = Categoria.objects.filter(nombre__iexact=nombre_custom).first()
                    if not cat_custom:
                        idx = 1
                        new_code = f"O{idx:02d}"
                        while Categoria.objects.filter(codigo=new_code).exists():
                            idx += 1
                            new_code = f"O{idx:02d}"
                        cat_custom = Categoria.objects.create(
                            nombre=nombre_custom,
                            codigo=new_code
                        )
                    reporte.categoria = cat_custom
                    reporte.titulo = f'Reporte de {cat_custom.nombre}'
                elif reporte.categoria:
                    reporte.titulo = f'Reporte de {reporte.categoria.nombre}'
                else:
                    reporte.titulo = 'Reporte ciudadano'

                reporte.save()

                # Guardar fotografías adicionales (2 y 3) como evidencias relacionales
                for foto_extra in fotos[1:3]:
                    Evidencia.objects.create(
                        reporte=reporte,
                        archivo=foto_extra,
                        descripcion='Evidencia fotográfica inicial del ciudadano'
                    )

                # Notificar a administradores y moderadores sobre la nueva incidencia registrada
                try:
                    admins_y_mods = User.objects.filter(
                        Q(is_superuser=True) | Q(perfilusuario__rol__in=['administrador', 'moderador'])
                    ).exclude(id=request.user.id).distinct()

                    for admin_mod in admins_y_mods:
                        crear_notificacion(
                            usuario=admin_mod,
                            reporte=reporte,
                            titulo='Nuevo reporte registrado',
                            mensaje=f'El usuario {request.user.username} ha registrado un nuevo reporte con folio {reporte.folio}.'
                        )
                except Exception as notif_err:
                    print(f"Error al enviar notificaciones: {notif_err}")

                try:
                    registrar_historial(
                        reporte=reporte,
                        usuario=request.user,
                        accion='Reporte registrado',
                        descripcion=f'El ciudadano generó un nuevo reporte en el sistema con {len(fotos)} fotografía(s) de evidencia.'
                    )
                except Exception as hist_err:
                    print(f"Error al registrar historial: {hist_err}")

                from django.urls import reverse
                return redirect(f"{reverse('detalle_reporte', args=[reporte.id])}?creado=1")

            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f'Ocurrió un error al guardar el reporte: {str(e)}')
                form.add_error(None, f'Error al guardar el reporte: {str(e)}')

    else:
        form = ReporteForm()

    return render(request, 'reportes/form_reporte.html', {
        'form': form,
        'mapbox_access_token': getattr(settings, 'MAPBOX_ACCESS_TOKEN', '')
    })


@login_required
def detalle_reporte(request, id):
    reporte = get_object_or_404(Reporte, id=id)
    rol = obtener_rol(request.user)

    if rol == 'ciudadano' and reporte.usuario != request.user:
        return redirect('panel_ciudadano')

    if request.user.is_authenticated:
        reporte.notificaciones.filter(usuario=request.user, leida=False).update(leida=True)

    evidencias = reporte.evidencias.all()

    return render(request, 'reportes/detalle_reporte.html', {
        'reporte': reporte,
        'rol': rol,
        'evidencias': evidencias,
    })


@rol_requerido('administrador', 'moderador', 'supervisor')
def cambiar_estado_reporte(request, id):
    reporte = get_object_or_404(Reporte, id=id)

    if request.method == 'POST':
        estado_anterior = reporte.estado
        nuevo_estado = request.POST.get('estado')

        reporte.estado = nuevo_estado
        reporte.save()

        crear_notificacion(
            usuario=reporte.usuario,
            reporte=reporte,
            titulo='Estado actualizado',
            mensaje=f'Tu reporte {reporte.folio} cambió de {estado_anterior} a {nuevo_estado}.'
        )

        registrar_historial(
            reporte=reporte,
            usuario=request.user,
            accion='Cambio de estado',
            descripcion=f'El estado cambió de {estado_anterior} a {nuevo_estado}.'
        )

        messages.success(request, 'Estado actualizado correctamente.')

    return redirect('detalle_reporte', id=reporte.id)


@rol_requerido('administrador', 'moderador', 'supervisor')
def cambiar_prioridad_reporte(request, id):
    reporte = get_object_or_404(Reporte, id=id)

    if request.method == 'POST':
        prioridad_anterior = reporte.prioridad
        nueva_prioridad = request.POST.get('prioridad')

        reporte.prioridad = nueva_prioridad
        reporte.save()

        crear_notificacion(
            usuario=reporte.usuario,
            reporte=reporte,
            titulo='Prioridad actualizada',
            mensaje=f'Tu reporte {reporte.folio} cambió de prioridad {prioridad_anterior} a {nueva_prioridad}.'
        )

        registrar_historial(
            reporte=reporte,
            usuario=request.user,
            accion='Cambio de prioridad',
            descripcion=f'La prioridad cambió de {prioridad_anterior} a {nueva_prioridad}.'
        )

        messages.success(request, 'Prioridad actualizada correctamente.')

    return redirect('detalle_reporte', id=reporte.id)


@rol_requerido('administrador', 'moderador', 'supervisor')
def actualizar_reporte(request, id):
    """
    Vista unificada para actualizar el estado, la prioridad y la evidencia de un reporte,
    generando una única notificación combinada para cambios técnicos y registrando en el historial.
    """
    reporte = get_object_or_404(Reporte, id=id)

    if request.method == 'POST':
        estado_anterior = reporte.estado
        nuevo_estado = request.POST.get('estado')
        prioridad_anterior = reporte.prioridad
        nueva_prioridad = request.POST.get('prioridad')

        cambios_msg = []
        historial_msg = []
        modificado = False

        if nuevo_estado and nuevo_estado != estado_anterior:
            reporte.estado = nuevo_estado
            cambios_msg.append(f"de estado ({estado_anterior} a {nuevo_estado})")
            historial_msg.append(f"El estado cambió de {estado_anterior} a {nuevo_estado}")
            modificado = True

        if nueva_prioridad and nueva_prioridad != prioridad_anterior:
            reporte.prioridad = nueva_prioridad
            cambios_msg.append(f"de prioridad ({prioridad_anterior} a {nueva_prioridad})")
            historial_msg.append(f"La prioridad cambió de {prioridad_anterior} a {nueva_prioridad}")
            modificado = True

        if modificado:
            reporte.save()

            # Enviar una sola notificación al ciudadano (usuario creador del reporte)
            mensaje_notif = f"Tu reporte {reporte.folio} se actualizó: " + " y ".join(cambios_msg) + "."
            crear_notificacion(
                usuario=reporte.usuario,
                reporte=reporte,
                titulo='Reporte actualizado',
                mensaje=mensaje_notif
            )

            # Registrar entradas en el historial
            for h_msg in historial_msg:
                registrar_historial(
                    reporte=reporte,
                    usuario=request.user,
                    accion='Reporte actualizado',
                    descripcion=h_msg
                )

            messages.success(request, 'Estado o prioridad actualizados correctamente.')

        # Procesar evidencia opcional si el usuario posee rol administrador/moderador (o es superuser)
        evidencia_creada = False
        rol = obtener_rol(request.user)
        if rol in ['administrador', 'moderador'] or request.user.is_superuser:
            archivo = request.FILES.get('archivo')
            descripcion_evidencia = request.POST.get('descripcion')
            if archivo:
                Evidencia.objects.create(
                    reporte=reporte,
                    archivo=archivo,
                    descripcion=descripcion_evidencia
                )
                evidencia_creada = True
                messages.success(request, 'Evidencia agregada correctamente.')

                registrar_historial(
                    reporte=reporte,
                    usuario=request.user,
                    accion='Evidencia agregada',
                    descripcion=f'Se subió una nueva imagen de evidencia: {descripcion_evidencia or "Sin descripción"}'
                )

        if not modificado and not evidencia_creada:
            messages.info(request, 'No se detectaron cambios.')

    return redirect('detalle_reporte', id=reporte.id)



@rol_requerido('administrador', 'moderador')
def panel_administrador(request):
    rol = obtener_rol(request.user)
    usuarios = User.objects.all().order_by('username')
    buscar_folio = request.GET.get('buscar_folio', '')

    reportes = Reporte.objects.all().order_by('-fecha_reporte')

    if buscar_folio:
       reportes = reportes.filter(folio__icontains=buscar_folio)
       
    categorias = Categoria.objects.all().order_by('nombre')

    puntos_mapa = []

    for reporte in reportes.exclude(latitud__isnull=True).exclude(longitud__isnull=True):
        puntos_mapa.append({
            'id': reporte.id,
            'folio': reporte.folio or 'Sin folio',
            'titulo': reporte.titulo or 'Reporte ciudadano',
            'categoria': reporte.categoria.nombre if reporte.categoria else 'Sin categoría',
            'estado': reporte.estado,
            'prioridad': reporte.prioridad,
            'colonia': reporte.colonia or 'No registrada',
            'latitud': float(reporte.latitud),
            'longitud': float(reporte.longitud),
            'url': f'/reporte/{reporte.id}/',
        })

    hoy = timezone.now()
    mes_actual = hoy.month
    anio_actual = hoy.year

    context = {
        'rol': rol,
        'usuarios': usuarios,
        'reportes': reportes,
        'categorias': categorias,
        'puntos_mapa': puntos_mapa,
        'buscar_folio': buscar_folio,
        'usuarios_registrados': usuarios.count(),
        'reportes_totales': reportes.count(),
        'reportes_pendientes': reportes.filter(estado='Pendiente').count(),
        'reportes_proceso': reportes.filter(estado='En proceso').count(),
        'reportes_resueltos': reportes.filter(estado='Resuelto').count(),
        'reportes_cancelados': reportes.filter(estado='Cancelado').count(),

        'total_baches': reportes.filter(categoria__nombre__icontains='Bache').count(),
        'total_basura': reportes.filter(categoria__nombre__icontains='Basura').count(),
        'mes_actual': mes_actual,
        'anio_actual': anio_actual,
    }

    return render(request, 'reportes/panel_administrador.html', context)



@rol_requerido('administrador')
def agregar_usuario(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        rol = request.POST.get('rol')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Ese nombre de usuario ya existe.')
            return redirect('panel_administrador')

        usuario = User.objects.create(
            username=username,
            email=email,
            password=make_password(password)
        )

        perfil, _ = PerfilUsuario.objects.get_or_create(usuario=usuario)
        perfil.rol = rol
        perfil.save()

        if rol == 'administrador':
            usuario.is_staff = True
            usuario.is_superuser = True

        usuario.save()

        messages.success(request, 'Usuario agregado correctamente.')

    return redirect('panel_administrador')


@rol_requerido('administrador')
def editar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)

    if request.method == 'POST':
        usuario.username = request.POST.get('username')
        usuario.email = request.POST.get('email')

        nueva_password = request.POST.get('password')
        if nueva_password:
            usuario.set_password(nueva_password)

        rol = request.POST.get('rol')

        perfil, _ = PerfilUsuario.objects.get_or_create(usuario=usuario)
        perfil.rol = rol
        perfil.save()

        if 'activo' in request.POST:
            usuario.is_active = request.POST.get('activo') == 'on'

        usuario.is_staff = rol in ['administrador', 'supervisor']
        usuario.is_superuser = rol == 'administrador'
        usuario.save()

        messages.success(request, 'Usuario actualizado correctamente.')

    return redirect('panel_administrador')


@rol_requerido('administrador')
def eliminar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)

    if usuario == request.user:
        messages.error(request, 'No puedes eliminar tu propio usuario.')
        return redirect('panel_administrador')

    usuario.delete()
    messages.success(request, 'Usuario eliminado correctamente.')

    return redirect('panel_administrador')



@rol_requerido('administrador')
def eliminar_reporte(request, id):
    reporte = get_object_or_404(Reporte, id=id)
    reporte.delete()

    messages.success(request, 'Reporte eliminado correctamente.')
    return redirect('panel_administrador')


@rol_requerido('administrador')
def agregar_catalogo(request, catalogo):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')

        if catalogo == 'categoria':
            Categoria.objects.get_or_create(
                nombre=nombre,
                defaults={'codigo': codigo}
            )

        messages.success(request, 'Catálogo actualizado correctamente.')

    return redirect('panel_administrador')


@rol_requerido('administrador', 'moderador')
def reporte_general(request):
    reportes = Reporte.objects.all().order_by('-fecha_reporte')

    sector = request.GET.get('sector')
    estado = request.GET.get('estado')
    prioridad = request.GET.get('prioridad')
    colonia = request.GET.get('colonia')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if sector:
        reportes = reportes.filter(categoria_id=sector)

    if estado:
        reportes = reportes.filter(estado=estado)

    if prioridad:
        reportes = reportes.filter(prioridad=prioridad)

    if colonia:
        reportes = reportes.filter(colonia__icontains=colonia)

    if fecha_inicio:
        reportes = reportes.filter(fecha_reporte__date__gte=fecha_inicio)

    if fecha_fin:
        reportes = reportes.filter(fecha_reporte__date__lte=fecha_fin)

    reportes_baches = reportes.filter(categoria__nombre__icontains='Bache')
    total_baches = reportes_baches.count()
    baches_pendientes = reportes_baches.filter(estado='Pendiente').count()
    baches_proceso = reportes_baches.filter(estado='En proceso').count()
    baches_resueltos = reportes_baches.filter(estado='Resuelto').count()

    reportes_basura = reportes.filter(categoria__nombre__icontains='Basura')
    total_basura = reportes_basura.count()
    basura_pendientes = reportes_basura.filter(estado='Pendiente').count()
    basura_proceso = reportes_basura.filter(estado='En proceso').count()
    basura_resueltos = reportes_basura.filter(estado='Resuelto').count()

    # Agrupación dinámica por categorías para todas las incidencias
    reportes_por_categoria = []
    for cat in Categoria.objects.all().order_by('nombre'):
        cat_reportes = reportes.filter(categoria=cat)
        reportes_por_categoria.append({
            'categoria': cat,
            'reportes': cat_reportes,
            'total': cat_reportes.count(),
            'pendientes': cat_reportes.filter(estado='Pendiente').count(),
            'en_proceso': cat_reportes.filter(estado='En proceso').count(),
            'resueltos': cat_reportes.filter(estado='Resuelto').count(),
        })

    context = {
        'reportes': reportes,
        'total_reportes': reportes.count(),
        'pendientes': reportes.filter(estado='Pendiente').count(),
        'en_proceso': reportes.filter(estado='En proceso').count(),
        'completados': reportes.filter(estado='Resuelto').count(),
        'cancelados': reportes.filter(estado='Cancelado').count(),

        'reportes_baches': reportes_baches,
        'total_baches': total_baches,
        'baches_pendientes': baches_pendientes,
        'baches_proceso': baches_proceso,
        'baches_resueltos': baches_resueltos,

        'reportes_basura': reportes_basura,
        'total_basura': total_basura,
        'basura_pendientes': basura_pendientes,
        'basura_proceso': basura_proceso,
        'basura_resueltos': basura_resueltos,

        'reportes_por_categoria': reportes_por_categoria,

        'categorias': Categoria.objects.all().order_by('nombre'),
        'colonias': Reporte.objects.exclude(colonia__isnull=True).exclude(colonia='').values_list('colonia', flat=True).distinct().order_by('colonia'),

        'sector_actual': sector or '',
        'estado_actual': estado or '',
        'prioridad_actual': prioridad or '',
        'colonia_actual': colonia or '',
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'todos_los_reportes': Reporte.objects.all().order_by('-fecha_reporte'),
    }

    return render(request, 'reportes/reporte_general.html', context)


@rol_requerido('administrador', 'moderador')
def agregar_evidencia(request, id):
    reporte = get_object_or_404(Reporte, id=id)

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        descripcion = request.POST.get('descripcion')

        if archivo:
            Evidencia.objects.create(
                reporte=reporte,
                archivo=archivo,
                descripcion=descripcion
            )

            messages.success(request, 'Evidencia agregada correctamente.')
        else:
            messages.error(request, 'Debes seleccionar una imagen.')

    return redirect('detalle_reporte', id=reporte.id)


@login_required
def notificaciones(request):
    lista = Notificacion.objects.filter(usuario=request.user).order_by('-fecha')
    return render(request, 'reportes/notificaciones.html', {
        'notificaciones': lista
    })


@login_required
def abrir_notificacion(request, id):
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    if not notificacion.leida:
        notificacion.leida = True
        notificacion.save()
    if notificacion.reporte:
        return redirect('detalle_reporte', id=notificacion.reporte.id)
    return redirect('notificaciones')


@login_required
def marcar_notificacion_leida_individual(request, id):
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    if not notificacion.leida:
        notificacion.leida = True
        notificacion.save()
        messages.success(request, 'Notificación marcada como leída.')
    return redirect('notificaciones')


@login_required
def eliminar_notificacion(request, id):
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    notificacion.delete()
    messages.success(request, 'Notificación eliminada correctamente.')
    return redirect('notificaciones')


@login_required
def marcar_notificaciones_leidas(request):
    Notificacion.objects.filter(
        usuario=request.user,
        leida=False
    ).update(leida=True)
    messages.success(request, 'Todas las notificaciones han sido marcadas como leídas.')
    return redirect('notificaciones')


@login_required
def api_notificaciones_sin_leer(request):
    """
    Retorna la cantidad de notificaciones sin leer del usuario autenticado en formato JSON.
    """
    unread_count = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    return JsonResponse({'unread_count': unread_count})

@rol_requerido('administrador')
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)

    reportes_asociados = Reporte.objects.filter(categoria=categoria).count()

    if reportes_asociados > 0:
        messages.error(
            request,
            f'No puedes eliminar esta categoría porque tiene {reportes_asociados} reporte(s) asociados.'
        )
        return redirect('panel_administrador')

    categoria.delete()
    messages.success(request, 'Categoría eliminada correctamente.')

    return redirect('panel_administrador')

@login_required
def reporte_individual_pdf(request, id):
    reporte = get_object_or_404(Reporte, id=id)

    rol = obtener_rol(request.user)

    if rol == 'ciudadano' and reporte.usuario != request.user:
        messages.error(request, 'No tienes permiso para descargar este reporte.')
        return redirect('panel_ciudadano')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{reporte.folio or reporte.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=35,
        leftMargin=35,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        'Titulo',
        parent=styles['Normal'],
        fontSize=13,
        alignment=1,
        fontName='Helvetica-Bold',
        leading=16
    )

    normal_style = ParagraphStyle(
        'NormalCustom',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )

    negrita_style = ParagraphStyle(
        'Negrita',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        fontName='Helvetica-Bold'
    )

    section_style = ParagraphStyle(
        'Section',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        fontName='Helvetica-Bold',
        textColor=colors.black
    )

    elementos = []

    # =============================
    # ENCABEZADO PRINCIPAL
    # =============================

    nombre_sistema = Paragraph(
        'SISTEMA DE REPORTES CIUDADANOS<br/>MUNICIPIO DE ACAPULCO DE JUÁREZ',
        titulo_style
    )

    datos_doc = Table([
        ['FOLIO', reporte.folio or 'Sin folio'],
        ['REV', '00'],
        ['FECHA ELAB.', reporte.fecha_reporte.strftime('%d/%m/%Y')],
    ], colWidths=[70, 90])

    datos_doc.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    # Cargar el logo de la empresa PIGRICA
    logo_path = os.path.join(settings.BASE_DIR, 'reportes', 'static', 'WhatsApp Image 2026-07-03 at 9.16.15 AM.jpeg')
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=32, height=32)
        logo_texto = Paragraph(
            '<b>PIGRICA</b><br/>ACAPULCO',
            ParagraphStyle(
                'LogoTexto',
                parent=styles['Normal'],
                fontSize=9,
                alignment=1,
                fontName='Helvetica-Bold'
            )
        )
        logo_display = Table([
            [logo_img, logo_texto]
        ], colWidths=[36, 68])
        logo_display.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
    else:
        logo_display = Paragraph(
            '<b>PIGRICA</b><br/>ACAPULCO',
            ParagraphStyle(
                'LogoTexto',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                fontName='Helvetica-Bold'
            )
        )

    encabezado = Table([
        [logo_display, nombre_sistema, datos_doc]
    ], colWidths=[120, 300, 160])

    encabezado.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(encabezado)
    elementos.append(Spacer(1, 12))

    # =============================
    # TABLA DATOS GENERALES
    # =============================

    nombre_ciudadano = reporte.usuario.get_full_name()
    if not nombre_ciudadano:
        nombre_ciudadano = reporte.usuario.username

    telefono = 'No registrado'
    if hasattr(reporte.usuario, 'perfilusuario'):
        telefono = reporte.usuario.perfilusuario.telefono or 'No registrado'

    datos_generales = [
        [
            Paragraph('<b>Tipo de problema:</b>', normal_style),
            str(reporte.categoria),
            Paragraph('<b>Fecha del reporte:</b>', normal_style),
            reporte.fecha_reporte.strftime('%d/%m/%Y')
        ],
        [
            Paragraph('<b>Sector:</b>', normal_style),
            str(reporte.categoria),
            Paragraph('<b>Estado:</b>', normal_style),
            reporte.estado
        ],
        [
            Paragraph('<b>Colonia:</b>', normal_style),
            reporte.colonia or 'No registrada',
            Paragraph('<b>Prioridad:</b>', normal_style),
            reporte.prioridad
        ],
    ]

    tabla_generales = Table(datos_generales, colWidths=[110, 190, 120, 160])
    tabla_generales.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elementos.append(tabla_generales)

    barra_1 = Table([[Paragraph('DATOS DEL CIUDADANO', section_style)]], colWidths=[580])
    barra_1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))

    elementos.append(barra_1)

    datos_ciudadano = [
        [
            Paragraph('<b>Nombre:</b>', normal_style),
            nombre_ciudadano,
            Paragraph('<b>Correo:</b>', normal_style),
            reporte.usuario.email or 'No registrado'
        ],
        [
            Paragraph('<b>Teléfono:</b>', normal_style),
            telefono,
            Paragraph('<b>Usuario:</b>', normal_style),
            reporte.usuario.username
        ],
    ]

    tabla_ciudadano = Table(datos_ciudadano, colWidths=[90, 200, 90, 200])
    tabla_ciudadano.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elementos.append(tabla_ciudadano)
    elementos.append(Spacer(1, 18))

    # =============================
    # DESCRIPCIÓN
    # =============================

    barra_descripcion = Table([[Paragraph('DESCRIPCIÓN DEL REPORTE', section_style)]], colWidths=[580])
    barra_descripcion.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    elementos.append(barra_descripcion)

    descripcion = reporte.descripcion or 'Sin descripción registrada.'

    descripcion_box = Table([
        [Paragraph(descripcion, normal_style)]
    ], colWidths=[580], rowHeights=[70])

    descripcion_box.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(descripcion_box)
    elementos.append(Spacer(1, 14))

    # =============================
    # UBICACIÓN
    # =============================

    barra_ubicacion = Table([[Paragraph('UBICACIÓN DEL REPORTE', section_style)]], colWidths=[580])
    barra_ubicacion.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    elementos.append(barra_ubicacion)

    datos_ubicacion = [
        [
            Paragraph('<b>Calle:</b>', normal_style),
            reporte.calle or 'No registrada',
            Paragraph('<b>Referencia:</b>', normal_style),
            reporte.referencia or 'No registrada'
        ],
        [
            Paragraph('<b>Latitud:</b>', normal_style),
            str(reporte.latitud) if reporte.latitud else 'No registrada',
            Paragraph('<b>Longitud:</b>', normal_style),
            str(reporte.longitud) if reporte.longitud else 'No registrada'
        ],
    ]

    tabla_ubicacion = Table(datos_ubicacion, colWidths=[80, 210, 90, 200])
    tabla_ubicacion.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elementos.append(tabla_ubicacion)
    elementos.append(Spacer(1, 14))

    # =============================
    # EVIDENCIA FOTOGRÁFICA
    # =============================

    barra_evidencia = Table([[Paragraph('EVIDENCIA FOTOGRÁFICA', section_style)]], colWidths=[580])
    barra_evidencia.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    elementos.append(barra_evidencia)

    imagenes = []

    if reporte.foto:
        try:
            if os.path.exists(reporte.foto.path):
                img = Image(reporte.foto.path)
                img.drawWidth = 160
                img.drawHeight = 120
                imagenes.append(img)
        except:
            pass

    for evidencia in reporte.evidencias.all()[:2]:
        try:
            if evidencia.archivo and os.path.exists(evidencia.archivo.path):
                img = Image(evidencia.archivo.path)
                img.drawWidth = 160
                img.drawHeight = 120
                imagenes.append(img)
        except:
            pass

    if imagenes:
        while len(imagenes) < 3:
            imagenes.append(Paragraph('Sin imagen', normal_style))

        tabla_imagenes = Table(
            [[imagenes[0], imagenes[1], imagenes[2]]],
            colWidths=[190, 190, 190],
            rowHeights=[135]
        )

        tabla_imagenes.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))

        elementos.append(tabla_imagenes)

    else:
        sin_imagen = Table(
            [[Paragraph('No se registró evidencia fotográfica.', normal_style)]],
            colWidths=[580],
            rowHeights=[80]
        )

        sin_imagen.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))

        elementos.append(sin_imagen)

    elementos.append(Spacer(1, 14))

    # =============================
    # SEGUIMIENTO
    # =============================

    ultima_observacion = reporte.historial_eventos.order_by('-fecha').first()

    observaciones = 'Sin observaciones registradas.'
    fecha_atencion = 'Sin fecha de atención.'

    if ultima_observacion:
        observaciones = ultima_observacion.descripcion or ultima_observacion.accion
        fecha_atencion = ultima_observacion.fecha.strftime('%d/%m/%Y %H:%M')

    barra_seguimiento = Table([[Paragraph('SEGUIMIENTO / OBSERVACIONES', section_style)]], colWidths=[580])
    barra_seguimiento.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    elementos.append(barra_seguimiento)

    datos_seguimiento = [
        [
            Paragraph('<b>Estado actual:</b>', normal_style),
            reporte.estado,
            Paragraph('<b>Fecha de atención:</b>', normal_style),
            fecha_atencion
        ],
        [
            Paragraph('<b>Observaciones del administrador:</b>', normal_style),
            Paragraph(observaciones, normal_style),
            '',
            ''
        ],
    ]

    tabla_seguimiento = Table(datos_seguimiento, colWidths=[150, 170, 120, 140])
    tabla_seguimiento.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('SPAN', (1, 1), (3, 1)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elementos.append(tabla_seguimiento)
    elementos.append(Spacer(1, 28))

    # =============================
    # FIRMAS
    # =============================

    firmas = Table([
        ['Elaboró:', 'Revisó:', 'Aprobó:'],
        ['', '', ''],
        ['_________________________', '_________________________', '_________________________']
    ], colWidths=[193, 193, 193], rowHeights=[18, 32, 22])

    firmas.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(firmas)

    pie = Table([
        ['REPORTE CIUDADANO', 'REV-00']
    ], colWidths=[480, 100])

    pie.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(pie)

    doc.build(elementos)
    return response

@rol_requerido('administrador', 'moderador')
def reporte_mensual_incidentes_pdf(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    sector = request.GET.get('sector')
    estado = request.GET.get('estado')
    prioridad = request.GET.get('prioridad')
    colonia = request.GET.get('colonia')
    foto = request.GET.get('foto')

    hoy = timezone.now()

    if mes:
        mes = int(mes)
    else:
        mes = hoy.month

    if anio:
        anio = int(anio)
    else:
        anio = hoy.year

    import datetime
    from django.utils.timezone import make_aware

    # Calcular el rango de fechas para el mes y año seleccionados.
    # Esto soluciona el problema de incompatibilidad de zonas horarias en MySQL en sistemas Windows,
    # donde las funciones de extracción directa (__month y __year) retornan vacío si no se han cargado
    # las tablas de zona horaria en el motor MySQL.
    fecha_inicio = datetime.datetime(anio, mes, 1, 0, 0, 0)
    
    if mes == 12:
        fecha_fin = datetime.datetime(anio + 1, 1, 1, 0, 0, 0)
    else:
        fecha_fin = datetime.datetime(anio, mes + 1, 1, 0, 0, 0)

    # Hacer consciente de la zona horaria (timezone-aware) si USE_TZ está habilitado
    if settings.USE_TZ:
        fecha_inicio = make_aware(fecha_inicio)
        fecha_fin = make_aware(fecha_fin)

    # Filtrar por rango para evitar delegar la extracción de fecha/mes a la base de datos
    reportes = Reporte.objects.filter(
        fecha_reporte__gte=fecha_inicio,
        fecha_reporte__lt=fecha_fin
    ).order_by('categoria__nombre', 'fecha_reporte')

    # Mensaje de depuración para ver en la consola del servidor local
    print(f"\n[DEBUG PDF] Mes seleccionado: {mes}, Año: {anio}")
    print(f"[DEBUG PDF] Rango de consulta: desde {fecha_inicio} hasta {fecha_fin}")
    print(f"[DEBUG PDF] Total de reportes encontrados en este rango: {reportes.count()}\n")

    # Filtro por incidente
    if sector == 'bache':
        reportes = reportes.filter(categoria__nombre__icontains='Bache')

    elif sector == 'basura':
        reportes = reportes.filter(categoria__nombre__icontains='Basura')

    elif sector:
        reportes = reportes.filter(categoria_id=sector)

    if estado:
        reportes = reportes.filter(estado=estado)

    if prioridad:
        reportes = reportes.filter(prioridad=prioridad)

    if colonia:
        reportes = reportes.filter(colonia__icontains=colonia)

    if foto == 'con_foto':
        reportes = reportes.exclude(foto='')

    elif foto == 'sin_foto':
        reportes = reportes.filter(foto='')

    MESES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    nombre_mes = MESES[mes]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_mensual_{mes}_{anio}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=11,
        leftMargin=11,
        topMargin=28,
        bottomMargin=28
    )

    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        'TituloPDF',
        parent=styles['Normal'],
        fontSize=15,
        alignment=1,
        fontName='Helvetica-Bold',
        leading=18
    )

    section_style = ParagraphStyle(
        'SectionPDF',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        fontName='Helvetica-Bold'
    )

    elementos = []

    # =========================
    # ENCABEZADO
    # =========================
    # Cargar el logo de la empresa PIGRICA
    logo_path = os.path.join(settings.BASE_DIR, 'reportes', 'static', 'WhatsApp Image 2026-07-03 at 9.16.15 AM.jpeg')
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=32, height=32)
        logo_texto = Paragraph(
            '<b>PIGRICA</b><br/>ACAPULCO',
            ParagraphStyle(
                'LogoTexto',
                parent=styles['Normal'],
                fontSize=9,
                alignment=1,
                fontName='Helvetica-Bold'
            )
        )
        encabezado_izq = Table([
            [logo_img, logo_texto]
        ], colWidths=[40, 90])
        encabezado_izq.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
    else:
        encabezado_izq = Paragraph(
            '<b>PIGRICA</b><br/>ACAPULCO',
            ParagraphStyle(
                'LogoTexto',
                parent=styles['Normal'],
                fontSize=9,
                alignment=1,
                fontName='Helvetica-Bold'
            )
        )

    encabezado_centro = Paragraph(
        f'SISTEMA DE REPORTES CIUDADANOS<br/>REPORTE MENSUAL DE INCIDENTES<br/>{nombre_mes} {anio}',
        titulo_style
    )

    encabezado_der = Table([
        ['FOLIO', f'RPT-{mes:02d}-{anio}'],
        ['REV', '00'],
        ['FECHA', hoy.strftime('%d/%m/%Y')],
    ], colWidths=[55, 90])

    encabezado_der.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    encabezado = Table([
        [encabezado_izq, encabezado_centro, encabezado_der]
    ], colWidths=[140, 470, 160])

    encabezado.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(encabezado)
    elementos.append(Spacer(1, 14))

    # =========================
    # RESUMEN GENERAL
    # =========================
    total_reportes = reportes.count()
    pendientes = reportes.filter(estado='Pendiente').count()
    en_proceso = reportes.filter(estado='En proceso').count()
    resueltos = reportes.filter(estado='Resuelto').count()
    cancelados = reportes.filter(estado='Cancelado').count()

    resumen = Table([
        ['RESUMEN GENERAL DEL PERIODO'],
        [
            f'Mes: {nombre_mes} {anio}',
            f'Total reportes: {total_reportes}',
            f'Pendientes: {pendientes}',
            f'En proceso: {en_proceso}',
            f'Resueltos: {resueltos}',
            f'Cancelados: {cancelados}',
        ],
    ], colWidths=[128, 128, 128, 128, 128, 128])

    resumen.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 7),
    ]))

    elementos.append(resumen)
    elementos.append(Spacer(1, 16))

    # =========================
    # AGRUPAR POR INCIDENTE
    # =========================
    categorias = {}

    for reporte in reportes:
        nombre = reporte.categoria.nombre if reporte.categoria else 'Sin categoría'

        if nombre not in categorias:
            categorias[nombre] = []

        categorias[nombre].append(reporte)

    contador = 1

    if not categorias:
        sin_datos = Table([
            ['NO SE ENCONTRARON REPORTES CON LOS FILTROS SELECCIONADOS']
        ], colWidths=[770])

        sin_datos.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 18),
        ]))

        elementos.append(sin_datos)

    for nombre_incidente, lista in categorias.items():
        barra = Table([
            [f'{contador}. INCIDENTE: {nombre_incidente.upper()}']
        ], colWidths=[770])

        barra.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))

        elementos.append(barra)

        datos = [
            ['Folio', 'Fecha', 'Sector', 'Tipo de Problema', 'Colonia', 'Estado', 'Prioridad']
        ]

        for r in lista:
            datos.append([
                r.folio or 'Sin folio',
                r.fecha_reporte.strftime('%d/%m/%Y'),
                str(r.categoria),
                r.titulo or str(r.categoria),
                r.colonia or 'No registrada',
                r.estado,
                r.prioridad,
            ])

        tabla = Table(
            datos,
            colWidths=[100, 80, 115, 160, 120, 100, 95],
            repeatRows=1
        )

        tabla.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))

        elementos.append(tabla)

        total_incidente = Table([
            [f'TOTAL {nombre_incidente.upper()}:', len(lista)]
        ], colWidths=[650, 120])

        total_incidente.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))

        elementos.append(total_incidente)
        elementos.append(Spacer(1, 14))

        contador += 1

    # =========================
    # RESUMEN FINAL
    # =========================
    elementos.append(Spacer(1, 16))

    resumen_final = Table([
        ['RESUMEN FINAL DEL PERIODO', ''],
        ['Total de reportes', total_reportes],
        ['Pendientes', pendientes],
        ['En proceso', en_proceso],
        ['Resueltos', resueltos],
        ['Cancelados', cancelados],
    ], colWidths=[580, 190])

    resumen_final.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elementos.append(resumen_final)
    elementos.append(Spacer(1, 24))

    firmas = Table([
        ['Elaboró:', 'Revisó:', 'Aprobó:'],
        ['', '', ''],
        ['_________________________', '_________________________', '_________________________'],
        ['Nombre y firma', 'Nombre y firma', 'Nombre y firma'],
    ], colWidths=[256, 256, 256], rowHeights=[18, 34, 20, 16])

    firmas.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(firmas)

    def agregar_pie(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.drawString(30, 18, 'REPORTE GENERADO POR EL SISTEMA')
        canvas.drawCentredString(400, 18, f'PÁGINA {canvas.getPageNumber()}')
        canvas.drawRightString(760, 18, f'RPT-{mes:02d}-{anio}')
        canvas.restoreState()

    doc.build(elementos, onFirstPage=agregar_pie, onLaterPages=agregar_pie)

    return response


@rol_requerido('administrador', 'moderador')
def reporte_mensual_incidentes_excel(request):
    """
    Genera un reporte analítico en formato Excel (.xlsx) con:
    1. Resumen ejecutivo de métricas clave.
    2. Tabla de porcentajes por Tipo de Incidencia + Gráfica de Pastel a color con porcentajes.
    3. Tabla de porcentajes por Estado del Reporte + Gráfica de Pastel a color con porcentajes.
    4. Hoja detallada con el registro completo de reportes filtrados.
    """
    import openpyxl
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime
    from django.utils.timezone import make_aware

    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    sector = request.GET.get('sector')
    estado = request.GET.get('estado')
    prioridad = request.GET.get('prioridad')
    colonia = request.GET.get('colonia')
    foto = request.GET.get('foto')

    hoy = timezone.now()
    mes = int(mes) if mes else hoy.month
    anio = int(anio) if anio else hoy.year

    fecha_inicio = datetime.datetime(anio, mes, 1, 0, 0, 0)
    if mes == 12:
        fecha_fin = datetime.datetime(anio + 1, 1, 1, 0, 0, 0)
    else:
        fecha_fin = datetime.datetime(anio, mes + 1, 1, 0, 0, 0)

    if settings.USE_TZ:
        fecha_inicio = make_aware(fecha_inicio)
        fecha_fin = make_aware(fecha_fin)

    reportes = Reporte.objects.filter(
        fecha_reporte__gte=fecha_inicio,
        fecha_reporte__lt=fecha_fin
    ).order_by('categoria__nombre', 'fecha_reporte')

    if sector == 'bache':
        reportes = reportes.filter(categoria__nombre__icontains='Bache')
    elif sector == 'basura':
        reportes = reportes.filter(categoria__nombre__icontains='Basura')
    elif sector:
        reportes = reportes.filter(categoria_id=sector)

    if estado:
        reportes = reportes.filter(estado=estado)
    if prioridad:
        reportes = reportes.filter(prioridad=prioridad)
    if colonia:
        reportes = reportes.filter(colonia__icontains=colonia)
    if foto == 'con_foto':
        reportes = reportes.exclude(foto='')
    elif foto == 'sin_foto':
        reportes = reportes.filter(foto='')

    MESES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    nombre_mes = MESES[mes]

    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # HOJA 1: RESUMEN Y GRÁFICAS (Dashboard Analítico)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Estadísticas y Gráficas"
    ws1.views.sheetView[0].showGridLines = True

    # Estilos principales
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    accent_orange = PatternFill(start_color="FF7417", end_color="FF7417", fill_type="solid")
    light_total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Arial", size=9.5, italic=True, color="FFFFFF")
    font_header = Font(name="Arial", size=10.5, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True, color="0F172A")
    font_regular = Font(name="Arial", size=10, color="334155")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='0F172A'),
        bottom=Side(style='double', color='0F172A')
    )

    # Banner de Encabezado ampliado (A1:L2)
    ws1.merge_cells("A1:L1")
    ws1.merge_cells("A2:L2")
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 20

    ws1["A1"] = "SISTEMA DE REPORTES CIUDADANOS - MUNICIPIO DE ACAPULCO DE JUÁREZ"
    ws1["A1"].font = font_title
    ws1["A1"].fill = header_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1["A2"] = f"REPORTE ANÁLITICO DE INCIDENCIAS ({nombre_mes} {anio})"
    ws1["A2"].font = Font(name="Arial", size=11, bold=True, color="F59E0B")
    ws1["A2"].fill = header_fill
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Fila de metadatos (A3:L3)
    ws1.merge_cells("A3:L3")
    ws1["A3"] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Total reportes analizados: {reportes.count()} | Operador: {request.user.username}"
    ws1["A3"].font = font_subtitle
    ws1["A3"].fill = sub_fill
    ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Separador
    ws1.row_dimensions[4].height = 12

    # -------------------------------------------------------------
    # TABLA 1: PORCENTAJE POR TIPO DE INCIDENCIA (CATEGORÍA)
    # -------------------------------------------------------------
    ws1["A5"] = "Tipo de Incidencia (Categoría)"
    ws1["B5"] = "Cantidad"
    ws1["C5"] = "Porcentaje"
    ws1.row_dimensions[5].height = 22
    for col in ["A5", "B5", "C5"]:
        ws1[col].font = font_header
        ws1[col].fill = accent_orange
        ws1[col].alignment = Alignment(horizontal="center", vertical="center")

    categorias = Categoria.objects.all().order_by('nombre')
    total_reps = reportes.count() or 1

    row_idx = 6
    cat_colors = {
        'baches': 'F59E0B',      # Naranja vial
        'basura': '10B981',      # Verde esmeralda
        'electricidad': 'EAB308',# Amarillo
        'agua': '06B6D4',        # Cian acua
        'alumbrado': '3B82F6',   # Azul rey
        'semaforo': 'EC4899',    # Rosa carmín
        'parques': '84CC16',     # Verde lima
        'seguridad': '8B5CF6'    # Púrpura
    }
    cat_slice_colors = []

    for cat in categorias:
        cant = reportes.filter(categoria=cat).count()
        if cant > 0:
            ws1.cell(row=row_idx, column=1, value=cat.nombre).font = font_regular
            ws1.cell(row=row_idx, column=2, value=cant).font = font_bold
            ws1.cell(row=row_idx, column=3, value=cant / total_reps).number_format = '0.0%'
            
            ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")

            for c in range(1, 4):
                ws1.cell(row=row_idx, column=c).border = thin_border

            cat_name_lower = cat.nombre.lower()
            color = '64748B'
            for key, val in cat_colors.items():
                if key in cat_name_lower:
                    color = val
                    break
            cat_slice_colors.append(color)

            row_idx += 1

    # Fila de Total para Tabla 1
    ws1.cell(row=row_idx, column=1, value="TOTAL").font = font_bold
    ws1.cell(row=row_idx, column=2, value=reportes.count()).font = font_bold
    ws1.cell(row=row_idx, column=3, value=1.0).number_format = '0.0%'
    ws1.cell(row=row_idx, column=3).font = font_bold

    for c in range(1, 4):
        cell = ws1.cell(row=row_idx, column=c)
        cell.fill = light_total_fill
        cell.border = double_bottom_border
        if c > 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    cat_data_end_row = row_idx - 1

    # Crear Gráfica de Pastel 1 (Categorías)
    if cat_data_end_row >= 6:
        pie_cat = PieChart()
        pie_cat.title = "Porcentaje por Tipo de Incidencia"
        if pie_cat.title:
            pie_cat.title.overlay = False
        pie_cat.width = 18
        pie_cat.height = 11.5
        if pie_cat.legend:
            pie_cat.legend.overlay = False
            pie_cat.legend.position = "r"

        labels_cat = Reference(ws1, min_col=1, min_row=6, max_row=cat_data_end_row)
        data_cat = Reference(ws1, min_col=2, min_row=5, max_row=cat_data_end_row)
        pie_cat.add_data(data_cat, titles_from_data=True)
        pie_cat.set_categories(labels_cat)

        if pie_cat.series:
            series = pie_cat.series[0]
            for idx, color in enumerate(cat_slice_colors):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = color
                series.data_points.append(dp)

        ws1.add_chart(pie_cat, "E5")

    # -------------------------------------------------------------
    # TABLA 2: PORCENTAJE POR ESTADO DEL REPORTE
    # -------------------------------------------------------------
    start_estado_row = 16
    ws1.cell(row=start_estado_row, column=1, value="Estado del Reporte").font = font_header
    ws1.cell(row=start_estado_row, column=2, value="Cantidad").font = font_header
    ws1.cell(row=start_estado_row, column=3, value="Porcentaje").font = font_header
    ws1.row_dimensions[start_estado_row].height = 22
    for c in range(1, 4):
        ws1.cell(row=start_estado_row, column=c).fill = header_fill
        ws1.cell(row=start_estado_row, column=c).alignment = Alignment(horizontal="center", vertical="center")

    estados_list = [
        ('Pendiente', reportes.filter(estado='Pendiente').count(), 'EF4444', 'FEE2E2', '991B1B'),
        ('En proceso', reportes.filter(estado='En proceso').count(), '3B82F6', 'DBEAFE', '1E40AF'),
        ('Resuelto', reportes.filter(estado='Resuelto').count(), '10B981', 'D1FAE5', '065F46'),
        ('Cancelado', reportes.filter(estado='Cancelado').count(), '64748B', 'F1F5F9', '475569'),
    ]

    e_row = start_estado_row + 1
    estado_slice_colors = []
    for nombre_e, cant_e, color_e, fill_bg, text_color in estados_list:
        ws1.cell(row=e_row, column=1, value=nombre_e).font = Font(name="Arial", size=10, bold=True, color=text_color)
        ws1.cell(row=e_row, column=2, value=cant_e).font = font_bold
        ws1.cell(row=e_row, column=3, value=cant_e / total_reps).number_format = '0.0%'
        ws1.cell(row=e_row, column=3).font = font_bold

        ws1.cell(row=e_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=e_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=e_row, column=3).alignment = Alignment(horizontal="center", vertical="center")

        row_fill = PatternFill(start_color=fill_bg, end_color=fill_bg, fill_type="solid")
        for c in range(1, 4):
            ws1.cell(row=e_row, column=c).border = thin_border
            ws1.cell(row=e_row, column=c).fill = row_fill

        estado_slice_colors.append(color_e)
        e_row += 1

    # Fila Total Tabla 2
    ws1.cell(row=e_row, column=1, value="TOTAL").font = font_bold
    ws1.cell(row=e_row, column=2, value=reportes.count()).font = font_bold
    ws1.cell(row=e_row, column=3, value=1.0).number_format = '0.0%'
    ws1.cell(row=e_row, column=3).font = font_bold

    for c in range(1, 4):
        cell = ws1.cell(row=e_row, column=c)
        cell.fill = light_total_fill
        cell.border = double_bottom_border
        if c > 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    e_data_end_row = e_row - 1

    # Crear Gráfica de Pastel 2 (Estados) posicionada en E27 (Sin solapamiento)
    pie_est = PieChart()
    pie_est.title = "Porcentaje por Estado del Reporte"
    if pie_est.title:
        pie_est.title.overlay = False
    pie_est.width = 18
    pie_est.height = 11.5
    if pie_est.legend:
        pie_est.legend.overlay = False
        pie_est.legend.position = "r"

    labels_est = Reference(ws1, min_col=1, min_row=start_estado_row + 1, max_row=e_data_end_row)
    data_est = Reference(ws1, min_col=2, min_row=start_estado_row, max_row=e_data_end_row)
    pie_est.add_data(data_est, titles_from_data=True)
    pie_est.set_categories(labels_est)

    if pie_est.series:
        series2 = pie_est.series[0]
        for idx, color in enumerate(estado_slice_colors):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            series2.data_points.append(dp)

    ws1.add_chart(pie_est, "E27")

    # Ajustar dimensiones de columnas para Hoja 1
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 3
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 16

    # -------------------------------------------------------------
    # HOJA 2: DETALLE REGISTRO COMPLETO DE REPORTES
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Registro de Incidencias")
    ws2.views.sheetView[0].showGridLines = True

    headers_ws2 = [
        "Folio", "Fecha", "Categoría / Incidencia", "Título", 
        "Descripción", "Ubicación (Calle/Colonia)", "Estado", "Prioridad", "Ciudadano"
    ]

    ws2.append(headers_ws2)
    for col_num in range(1, len(headers_ws2) + 1):
        c = ws2.cell(row=1, column=col_num)
        c.font = font_header
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[1].height = 25

    for rep in reportes:
        ubicacion = f"{rep.calle or ''}, {rep.colonia or ''}".strip(', ')
        ws2.append([
            rep.folio or f"REP-{rep.id}",
            rep.fecha_reporte.strftime("%d/%m/%Y %H:%M"),
            rep.categoria.nombre if rep.categoria else "Sin categoría",
            rep.titulo or "",
            rep.descripcion or "",
            ubicacion or "No especificada",
            rep.estado,
            rep.prioridad,
            rep.usuario.username
        ])

    for row in range(2, ws2.max_row + 1):
        for col in range(1, len(headers_ws2) + 1):
            cell = ws2.cell(row=row, column=col)
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_incidentes_{mes}_{anio}.xlsx"'
    wb.save(response)
    return response

from django.shortcuts import render
from django.db.models import Count
from .models import Reporte

def inicio(request):
    reportes = Reporte.objects.all()
    
    pendientes = reportes.filter(estado='Pendiente').count()
    proceso = reportes.filter(estado='En proceso').count()
    resueltos = reportes.filter(estado='Resuelto').count()
    total = reportes.count()

    # Excluimos reportes sin ubicación y los cancelados
    reportes_con_ubicacion = reportes.exclude(latitud__isnull=True).exclude(longitud__isnull=True).exclude(estado='Cancelado')
    
    puntos_mapa = []
    for rep in reportes_con_ubicacion:
        if rep.latitud and rep.longitud:
            # Asumiendo que tu modelo Categoría tiene un campo de texto (por ejemplo 'nombre' o 'titulo' o __str__)
            # Si es un campo de texto directo, usa rep.categoria; si es ForeignKey, usa rep.categoria.nombre
            try:
                categoria_texto = rep.categoria.nombre  # Cambiar por .titulo o dejar solo rep.categoria si es CharField
            except AttributeError:
                categoria_texto = str(rep.categoria)

            puntos_mapa.append({
                'latitud': float(rep.latitud),
                'longitud': float(rep.longitud),
                'categoria': categoria_texto,
                'estado': rep.estado,
            })

    resenas = Resena.objects.all().order_by('-fecha')[:6]

    reportes_resueltos_usuario = []
    has_resueltos_evaluados = False
    if request.user.is_authenticated:
        # Filtrar solo los reportes resueltos del usuario que AÚN NO tienen una reseña registrada
        reportes_resueltos_usuario = Reporte.objects.filter(
            usuario=request.user,
            estado='Resuelto',
            reseñas__isnull=True
        ).order_by('-fecha_reporte')

        # Saber si el usuario ya evaluó reportes resueltos anteriormente
        has_resueltos_evaluados = Reporte.objects.filter(
            usuario=request.user,
            estado='Resuelto',
            reseñas__isnull=False
        ).exists()

    context = {
        'reportes': reportes,
        'pendientes': pendientes,
        'proceso': proceso,
        'resueltos': resueltos,
        'total': total,
        # Mostrar en la galería los reportes resueltos que tengan al menos una foto y estén aprobados para mostrarse
        'resueltas_lista': reportes.filter(estado='Resuelto', mostrar_en_galeria=True).filter(
            Q(evidencias__isnull=False) | (~Q(foto='') & ~Q(foto__isnull=True))
        ).distinct(),
        'puntos_mapa': puntos_mapa,
        'resenas': resenas,
        'reportes_resueltos_usuario': reportes_resueltos_usuario,
        'has_resueltos_evaluados': has_resueltos_evaluados,
    }
    
    return render(request, "reportes/inicio.html", context)


@login_required
def dejar_reseña(request):
    if request.method == 'POST':
        puntuacion = request.POST.get('puntuacion', 5)
        comentario = request.POST.get('comentario', '').strip()
        reporte_id = request.POST.get('reporte_id')
        
        try:
            puntuacion = int(puntuacion)
            if puntuacion < 1 or puntuacion > 5:
                puntuacion = 5
        except (ValueError, TypeError):
            puntuacion = 5

        if not comentario:
            messages.error(request, 'El comentario de la reseña no puede estar vacío.')
            return redirect('inicio')

        reporte_asociado = None
        if reporte_id:
            reporte_asociado = get_object_or_404(Reporte, id=reporte_id, usuario=request.user, estado='Resuelto')
            # Validar que no exista previamente una reseña para este reporte resuelto
            if Resena.objects.filter(reporte=reporte_asociado).exists():
                messages.error(request, f'Ya has registrado una reseña previamente para el reporte {reporte_asociado.folio}.')
                return redirect('inicio')

        # Capturar fotografías adjuntas a la reseña
        fotos = request.FILES.getlist('fotos_reseña')
        if not fotos and 'fotos_reseña' in request.FILES:
            fotos = [request.FILES['fotos_reseña']]

        # Si el usuario seleccionó un reporte resuelto o intenta subir fotografías
        if reporte_asociado or fotos:
            if len(fotos) < 1:
                messages.error(request, 'Es obligatorio adjuntar al menos 1 fotografía JPG como evidencia al evaluar tu reporte resuelto.')
                return redirect('inicio')
            if len(fotos) > 2:
                messages.error(request, 'Únicamente se permite un máximo de 2 fotografías para la reseña.')
                return redirect('inicio')

            for f in fotos:
                ext = f.name.split('.')[-1].lower() if '.' in f.name else ''
                if ext not in ['jpg', 'jpeg']:
                    messages.error(request, f'La imagen "{f.name}" no está en formato JPG (.jpg / .jpeg).')
                    return redirect('inicio')

        resena = Resena.objects.create(
            usuario=request.user,
            reporte=reporte_asociado,
            comentario=comentario,
            puntuacion=puntuacion,
            foto1=fotos[0] if len(fotos) >= 1 else None,
            foto2=fotos[1] if len(fotos) >= 2 else None
        )

        messages.success(request, '¡Gracias por compartir tu reseña y evidencia del reporte resuelto!')

    return redirect('inicio')


@rol_requerido('administrador', 'moderador')
def editar_fotos_galeria(request, id):
    """
    Permite al administrador modificar las fotos de la galería pública ("Antes" y "Después")
    de un reporte que ya está en estado 'Resuelto'.
    """
    reporte = get_object_or_404(Reporte, id=id, estado='Resuelto')

    if request.method == 'POST':
        foto_antes = request.FILES.get('foto_antes')
        foto_despues = request.FILES.get('foto_despues')

        if foto_antes:
            reporte.foto = foto_antes
            reporte.save()
            messages.success(request, f'Foto "Antes" del reporte {reporte.folio} actualizada con éxito.')

        if foto_despues:
            evidencia = reporte.evidencias.first()
            if evidencia:
                evidencia.archivo = foto_despues
                evidencia.save()
            else:
                Evidencia.objects.create(
                    reporte=reporte,
                    archivo=foto_despues,
                    descripcion='Evidencia de reparación'
                )
            messages.success(request, f'Foto "Después" del reporte {reporte.folio} actualizada con éxito.')

        if not foto_antes and not foto_despues:
            messages.info(request, 'No se seleccionó ninguna imagen para subir.')

    return redirect('panel_administrador')


@rol_requerido('administrador', 'moderador')
def eliminar_fotos_galeria(request, id):
    """
    Permite al administrador limpiar las fotos de la galería de un reporte resuelto,
    borrando tanto el 'Antes' como el 'Después' (evidencias).
    """
    reporte = get_object_or_404(Reporte, id=id, estado='Resuelto')

    # Limpiar foto Antes
    if reporte.foto:
        try:
            reporte.foto.delete(save=False)
        except Exception:
            pass
        reporte.foto = None
        reporte.save()

    # Limpiar foto Después (Evidencias)
    reporte.evidencias.all().delete()

    messages.success(request, f'Fotos de la galería para el reporte {reporte.folio} eliminadas correctamente.')
    return redirect('panel_administrador')


@rol_requerido('administrador', 'moderador')
def eliminar_resena(request, id):
    """
    Permite a administradores o moderadores eliminar una reseña inapropiada.
    """
    resena = get_object_or_404(Resena, id=id)
    resena.delete()
    messages.success(request, 'Reseña eliminada correctamente.')
    return redirect('inicio')


@rol_requerido('administrador')
def toggle_bloqueo_usuario(request, id):
    """
    Permite al administrador activar/desactivar (bloquear/desbloquear) la cuenta de un usuario.
    """
    usuario = get_object_or_404(User, id=id)
    if usuario == request.user:
        messages.error(request, 'No puedes bloquear o desbloquear tu propia cuenta.')
    else:
        usuario.is_active = not usuario.is_active
        usuario.save()
        estado = 'bloqueado' if not usuario.is_active else 'activado'
        messages.success(request, f'El usuario {usuario.username} ha sido {estado} correctamente.')
    return redirect('panel_administrador')


@rol_requerido('administrador', 'moderador')
def toggle_mostrar_galeria(request, id):
    """
    Permite a administradores y moderadores decidir si mostrar (subir) o no mostrar (quitar)
    una incidencia resuelta en la galería pública de la landing page.
    """
    reporte = get_object_or_404(Reporte, id=id, estado='Resuelto')
    reporte.mostrar_en_galeria = not reporte.mostrar_en_galeria
    reporte.save()

    estado_texto = 'visible en' if reporte.mostrar_en_galeria else 'ocultado de'
    messages.success(request, f'El reporte {reporte.folio} ahora está {estado_texto} la galería pública.')
    return redirect('panel_administrador')


def custom_csrf_failure(request, reason=""):
    """
    Vista personalizada para fallos de verificación CSRF.
    Muestra un mensaje de error amigable en español y protege los datos sensibles (ocultando contraseñas).
    """
    return render(request, 'reportes/csrf_error.html', {
        'reason': reason,
        'path': request.path
    }, status=403)


@login_required
def enviar_contacto_correo(request):
    """
    Procesa el formulario de contacto para enviar un correo mediante SMTP de Gmail.
    """
    if request.method == 'POST':
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        
        usuario_actual = request.user
        
        cuerpo_correo = f"""
        Has recibido un nuevo mensaje de contacto desde tu plataforma web.
        
        Detalles del remitente:
        - Usuario: {usuario_actual.username}
        - Correo registrado: {usuario_actual.email}
        
        Mensaje:
        {mensaje}
        """
        
        try:
            send_mail(
                subject=f"Contacto Web: {asunto}",
                message=cuerpo_correo,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[settings.EMAIL_HOST_USER],
                fail_silently=False,
            )
            return JsonResponse({'status': 'success', 'message': 'Mensaje enviado correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error al enviar: {str(e)}'})
            
    return redirect('inicio')